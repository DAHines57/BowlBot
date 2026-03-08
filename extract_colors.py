"""
One-time script to extract team colors from Bowling-Friends League v5.xlsx
and write them to team_colors.json.

Run whenever team names or colors change:
    venv\Scripts\python.exe extract_colors.py
"""
import json
import os
from openpyxl import load_workbook

EXCEL_FILE = "Bowling-Friends League v5.xlsx"
OUTPUT_FILE = "team_colors.json"

# Column indices (1-based): Team=2, Player=3
TEAM_COL = 2


def _get_color(cell):
    """Return a 6-digit hex color string from a cell's fill, or None."""
    ft = cell.fill.fill_type
    if not ft or ft == "none":
        return None
    fg = cell.fill.fgColor
    if fg.type == "rgb":
        argb = fg.rgb  # e.g. "FF4F81BD"
        # Strip alpha channel if present
        return argb[-6:] if len(argb) == 8 else argb
    return None


def extract_team_colors(path: str) -> dict:
    """Scan all Season sheets and return {team_name: hex_color}."""
    wb = load_workbook(path, data_only=False)
    colors = {}  # team_name -> hex color (most recent season wins)

    # Sort sheets newest-first so most recent color takes precedence
    season_sheets = sorted(
        [s for s in wb.sheetnames if s.startswith("Season") and s.split()[-1].isdigit()],
        key=lambda s: int(s.split()[-1]),
        reverse=True,
    )

    for sheet_name in season_sheets:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            team_cell = ws.cell(row=row, column=TEAM_COL)
            team_name = team_cell.value
            if not team_name or not isinstance(team_name, str):
                continue
            team_name = team_name.strip()
            if team_name in colors:
                continue  # already captured from a newer season
            color = _get_color(team_cell)
            if color:
                colors[team_name] = f"#{color}"

    return colors


if __name__ == "__main__":
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: {EXCEL_FILE} not found.")
        raise SystemExit(1)

    colors = extract_team_colors(EXCEL_FILE)
    with open(OUTPUT_FILE, "w") as f:
        json.dump(colors, f, indent=2)

    print(f"Wrote {len(colors)} team colors to {OUTPUT_FILE}:")
    for team, color in sorted(colors.items()):
        print(f"  {color}  {team}")
