"""Extract team name colors from Excel workbook cell fills."""
from __future__ import annotations

from openpyxl import load_workbook

from stats.compute import parse_season_number
from stats.facts import canonical_team_name

TEAM_COL = 2


def cell_fill_hex(cell) -> str | None:
    """Return #RRGGBB from a cell fill, or None."""
    ft = cell.fill.fill_type
    if not ft or ft == "none":
        return None
    fg = cell.fill.fgColor
    if fg.type == "rgb" and fg.rgb:
        argb = str(fg.rgb)
        hex6 = argb[-6:] if len(argb) == 8 else argb
        return f"#{hex6.upper()}"
    return None


def extract_team_colors_for_sheet(file_path: str, sheet_key: str) -> dict[str, str]:
    """{canonical_team_name: #hex} from one season sheet."""
    wb = load_workbook(file_path, data_only=False)
    try:
        if sheet_key not in wb.sheetnames:
            return {}
        ws = wb[sheet_key]
        colors: dict[str, str] = {}
        for row in range(2, ws.max_row + 1):
            team_cell = ws.cell(row=row, column=TEAM_COL)
            raw = team_cell.value
            if not raw or not isinstance(raw, str):
                continue
            sn = parse_season_number(sheet_key)
            name = canonical_team_name(raw.strip(), season_num=sn)
            if name in colors:
                continue
            hex_c = cell_fill_hex(team_cell)
            if hex_c:
                colors[name] = hex_c
        return colors
    finally:
        wb.close()


def build_season_color_maps(
    file_path: str, sheet_keys: list[str]
) -> dict[str, dict[str, str]]:
    """Per-season-sheet color maps in one workbook pass."""
    wb = load_workbook(file_path, data_only=False)
    try:
        out: dict[str, dict[str, str]] = {}
        for sheet_key in sheet_keys:
            if sheet_key not in wb.sheetnames:
                out[sheet_key] = {}
                continue
            ws = wb[sheet_key]
            colors: dict[str, str] = {}
            for row in range(2, ws.max_row + 1):
                team_cell = ws.cell(row=row, column=TEAM_COL)
                raw = team_cell.value
                if not raw or not isinstance(raw, str):
                    continue
                sn = parse_season_number(sheet_key)
                name = canonical_team_name(raw.strip(), season_num=sn)
                if name in colors:
                    continue
                hex_c = cell_fill_hex(team_cell)
                if hex_c:
                    colors[name] = hex_c
            out[sheet_key] = colors
        return out
    finally:
        wb.close()


def extract_all_team_colors(file_path: str) -> dict[str, str]:
    """Global map; newest season sheet wins (matches legacy extract_colors.py)."""
    wb = load_workbook(file_path, data_only=False)
    try:
        season_sheets = sorted(
            [
                s
                for s in wb.sheetnames
                if s.startswith("Season") and s.split()[-1].isdigit()
            ],
            key=lambda s: int(s.split()[-1]),
            reverse=True,
        )
        colors: dict[str, str] = {}
        for sheet_name in season_sheets:
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                team_cell = ws.cell(row=row, column=TEAM_COL)
                raw = team_cell.value
                if not raw or not isinstance(raw, str):
                    continue
                sn = parse_season_number(sheet_name)
                name = canonical_team_name(raw.strip(), season_num=sn)
                if name in colors:
                    continue
                hex_c = cell_fill_hex(team_cell)
                if hex_c:
                    colors[name] = hex_c
        return colors
    finally:
        wb.close()
