"""
Migration script to convert old Excel structure to new database-like structure.
Handles both detailed seasons (with individual games) and aggregated seasons.
"""
import openpyxl
from openpyxl import load_workbook
from collections import defaultdict
from typing import Dict, List, Set, Optional
import re


class ExcelMigrator:
    """Migrates data from old Excel format to new database structure."""
    
    def __init__(self, old_file: str, new_file: str):
        self.old_wb = load_workbook(old_file, data_only=True)
        self.new_wb = load_workbook(new_file)
        self.players_map = {}  # player_name -> player_id
        self.teams_map = {}  # (season_id, team_name) -> team_id
        self.seasons_map = {}  # season_name -> season_id
        self.next_player_id = 1
        self.next_team_id = 1
        self.next_season_id = 1
        self.next_game_id = 1
        self.next_aggregate_id = 1
        self.next_standing_id = 1
        self.next_team_player_id = 1
        
    def _safe_float(self, value, default=0.0):
        """Safely convert to float."""
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            try:
                return float(value)
            except (ValueError, TypeError):
                return default
        return default
    
    def _safe_int(self, value, default=0):
        """Safely convert to int."""
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str):
            try:
                return int(float(value))
            except (ValueError, TypeError):
                return default
        return default
    
    def migrate_seasons(self):
        """Extract and migrate season information."""
        print("Migrating seasons...")
        ws_seasons = self.new_wb["seasons"]
        
        # Clear example data (keep header)
        for row in range(ws_seasons.max_row, 1, -1):
            ws_seasons.delete_rows(row)
        
        season_sheets = [s for s in self.old_wb.sheetnames if s.startswith('Season')]
        season_sheets.sort(key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 0)
        
        for season_name in season_sheets:
            season_num = int(season_name.split()[-1])
            season_id = self.next_season_id
            self.seasons_map[season_name] = season_id
            
            # Determine if active (assume highest number is active)
            is_active = (season_num == max([int(s.split()[-1]) for s in season_sheets if s.split()[-1].isdigit()]))
            
            ws_seasons.append([
                season_id,
                season_num,
                season_name,
                "",  # start_date
                "",  # end_date
                "TRUE" if is_active else "FALSE",
                ""  # created_at
            ])
            
            self.next_season_id += 1
            print(f"  Added {season_name} (ID: {season_id})")
    
    def migrate_players(self):
        """Extract unique players from all seasons."""
        print("Migrating players...")
        ws_players = self.new_wb["players"]
        
        # Clear example data
        for row in range(ws_players.max_row, 1, -1):
            ws_players.delete_rows(row)
        
        players_set = set()
        season_sheets = [s for s in self.old_wb.sheetnames if s.startswith('Season')]
        
        # Extract players from each season
        for season_name in season_sheets:
            sheet = self.old_wb[season_name]
            start_row = 13
            
            for row in range(start_row, sheet.max_row + 1):
                player_cell = sheet.cell(row=row, column=3).value
                if player_cell and isinstance(player_cell, str) and player_cell.strip():
                    player_name = player_cell.strip()
                    if player_name.lower() not in ['player', 'team averages', 'average']:
                        players_set.add(player_name)
        
        # Add players to sheet
        for player_name in sorted(players_set):
            player_id = self.next_player_id
            self.players_map[player_name] = player_id
            
            ws_players.append([
                player_id,
                player_name,
                "",  # created_at
                ""   # updated_at
            ])
            
            self.next_player_id += 1
        
        print(f"  Found {len(players_set)} unique players")
    
    def migrate_teams(self):
        """Extract teams from each season."""
        print("Migrating teams...")
        ws_teams = self.new_wb["teams"]
        
        # Clear example data
        for row in range(ws_teams.max_row, 1, -1):
            ws_teams.delete_rows(row)
        
        season_sheets = [s for s in self.old_wb.sheetnames if s.startswith('Season')]
        
        for season_name in season_sheets:
            if season_name not in self.seasons_map:
                continue
            
            season_id = self.seasons_map[season_name]
            sheet = self.old_wb[season_name]
            start_row = 2
            teams_found = set()
            
            # Extract teams from standings section
            for row in range(start_row, min(start_row + 15, sheet.max_row + 1)):
                team_cell = sheet.cell(row=row, column=2).value
                if team_cell and isinstance(team_cell, str) and team_cell.strip():
                    team_name = team_cell.strip()
                    if team_name.lower() not in ['team', 'wins', 'losses', 'ties']:
                        teams_found.add(team_name)
            
            # Also check player section for teams
            for row in range(13, sheet.max_row + 1):
                team_cell = sheet.cell(row=row, column=2).value
                if team_cell and isinstance(team_cell, str) and team_cell.strip():
                    team_name = team_cell.strip()
                    if team_name.lower() not in ['team', 'player', 'team averages']:
                        teams_found.add(team_name)
            
            # Add teams to sheet
            for team_name in sorted(teams_found):
                team_id = self.next_team_id
                self.teams_map[(season_id, team_name)] = team_id
                
                ws_teams.append([
                    team_id,
                    season_id,
                    team_name,
                    ""  # created_at
                ])
                
                self.next_team_id += 1
            
            print(f"  {season_name}: {len(teams_found)} teams")
    
    def migrate_team_players(self):
        """Link players to teams for each season."""
        print("Migrating team-player relationships...")
        ws_team_players = self.new_wb["team_players"]
        
        # Clear example data
        for row in range(ws_team_players.max_row, 1, -1):
            ws_team_players.delete_rows(row)
        
        season_sheets = [s for s in self.old_wb.sheetnames if s.startswith('Season')]
        
        for season_name in season_sheets:
            if season_name not in self.seasons_map:
                continue
            
            season_id = self.seasons_map[season_name]
            sheet = self.old_wb[season_name]
            current_team = None
            current_team_id = None
            
            # Extract from player scores section
            for row in range(13, sheet.max_row + 1):
                team_cell = sheet.cell(row=row, column=2).value
                player_cell = sheet.cell(row=row, column=3).value
                
                # Update current team
                if team_cell and isinstance(team_cell, str) and team_cell.strip():
                    team_name = team_cell.strip()
                    if team_name.lower() not in ['team', 'player', 'team averages']:
                        current_team = team_name
                        current_team_id = self.teams_map.get((season_id, team_name))
                
                # Add player-team relationship
                if player_cell and isinstance(player_cell, str) and player_cell.strip():
                    player_name = player_cell.strip()
                    if (player_name.lower() not in ['player', 'team averages', 'average'] and 
                        player_name in self.players_map and current_team_id):
                        
                        player_id = self.players_map[player_name]
                        
                        ws_team_players.append([
                            self.next_team_player_id,
                            current_team_id,
                            player_id,
                            season_id,
                            "FALSE",  # is_captain (could be enhanced)
                            "",  # draft_round
                            ""  # created_at
                        ])
                        
                        self.next_team_player_id += 1
            
            print(f"  {season_name}: Linked players to teams")
    
    def migrate_games_and_aggregates(self, detailed_seasons_start=10):
        """Migrate individual games for detailed seasons, aggregates for older seasons."""
        print("Migrating game data...")
        ws_games = self.new_wb["games"]
        ws_aggregates = self.new_wb["player_aggregates"]
        
        # Clear example data
        for row in range(ws_games.max_row, 1, -1):
            ws_games.delete_rows(row)
        for row in range(ws_aggregates.max_row, 1, -1):
            ws_aggregates.delete_rows(row)
        
        season_sheets = [s for s in self.old_wb.sheetnames if s.startswith('Season')]
        season_sheets.sort(key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 0)
        
        for season_name in season_sheets:
            if season_name not in self.seasons_map:
                continue
            
            season_id = self.seasons_map[season_name]
            season_num = int(season_name.split()[-1])
            sheet = self.old_wb[season_name]
            
            # Determine if this season has detailed game data
            has_detailed_games = season_num >= detailed_seasons_start
            
            current_team = None
            current_team_id = None
            
            # Extract player data
            for row in range(13, sheet.max_row + 1):
                team_cell = sheet.cell(row=row, column=2).value
                player_cell = sheet.cell(row=row, column=3).value
                
                # Update current team
                if team_cell and isinstance(team_cell, str) and team_cell.strip():
                    team_name = team_cell.strip()
                    if team_name.lower() not in ['team', 'player', 'team averages']:
                        current_team = team_name
                        current_team_id = self.teams_map.get((season_id, team_name))
                
                # Process player
                if player_cell and isinstance(player_cell, str) and player_cell.strip():
                    player_name = player_cell.strip()
                    if (player_name.lower() not in ['player', 'team averages', 'average'] and 
                        player_name in self.players_map and current_team_id):
                        
                        player_id = self.players_map[player_name]
                        
                        # Extract scores from week columns (starting at column 4)
                        scores = []
                        week_scores = {}
                        
                        for col in range(4, min(20, sheet.max_column + 1)):
                            score = sheet.cell(row=row, column=col).value
                            if score is not None:
                                score_float = self._safe_float(score)
                                if score_float > 0:  # Valid score
                                    week_num = col - 3  # Column 4 = Week 1
                                    scores.append((week_num, score_float))
                                    week_scores[week_num] = score_float
                        
                        if has_detailed_games and scores:
                            # Create individual game records
                            for week_num, score in scores:
                                ws_games.append([
                                    self.next_game_id,
                                    player_id,
                                    current_team_id,
                                    season_id,
                                    week_num,
                                    int(score),
                                    "",  # game_date
                                    "FALSE",  # is_playoff
                                    ""  # created_at
                                ])
                                self.next_game_id += 1
                        else:
                            # Create aggregate record
                            total_games = len(scores) if scores else 0
                            total_pins = sum(score for _, score in scores)
                            average = total_pins / total_games if total_games > 0 else 0
                            high_game = max((score for _, score in scores), default=0)
                            low_game = min((score for _, score in scores), default=0)
                            
                            # Build week columns
                            week_data = [""] * 7
                            for week_num, score in week_scores.items():
                                if 1 <= week_num <= 7:
                                    week_data[week_num - 1] = int(score)
                            
                            ws_aggregates.append([
                                self.next_aggregate_id,
                                player_id,
                                current_team_id,
                                season_id,
                                total_games,
                                int(total_pins),
                                round(average, 2) if average > 0 else "",
                                int(high_game) if high_game > 0 else "",
                                int(low_game) if low_game > 0 else "",
                                *week_data,  # week_1 through week_7
                                f"{season_name} - {'Aggregated data' if not scores else 'Partial data'}",
                                ""  # created_at
                            ])
                            
                            self.next_aggregate_id += 1
            
            if has_detailed_games:
                print(f"  {season_name}: Migrated individual games")
            else:
                print(f"  {season_name}: Migrated as aggregates")
    
    def migrate_team_standings(self):
        """Migrate team standings."""
        print("Migrating team standings...")
        ws_standings = self.new_wb["team_standings"]
        
        # Clear example data
        for row in range(ws_standings.max_row, 1, -1):
            ws_standings.delete_rows(row)
        
        season_sheets = [s for s in self.old_wb.sheetnames if s.startswith('Season')]
        
        for season_name in season_sheets:
            if season_name not in self.seasons_map:
                continue
            
            season_id = self.seasons_map[season_name]
            sheet = self.old_wb[season_name]
            start_row = 2
            
            # Extract standings
            for row in range(start_row, min(start_row + 15, sheet.max_row + 1)):
                team_cell = sheet.cell(row=row, column=2).value
                if team_cell and isinstance(team_cell, str) and team_cell.strip():
                    team_name = team_cell.strip()
                    if team_name.lower() not in ['team', 'wins', 'losses', 'ties']:
                        team_id = self.teams_map.get((season_id, team_name))
                        if team_id:
                            wins = self._safe_int(sheet.cell(row=row, column=3).value)
                            losses = self._safe_int(sheet.cell(row=row, column=4).value)
                            ties = self._safe_int(sheet.cell(row=row, column=5).value)
                            pins_for = self._safe_int(sheet.cell(row=row, column=6).value)
                            pins_against = self._safe_int(sheet.cell(row=row, column=7).value)
                            avg_per_game = self._safe_float(sheet.cell(row=row, column=8).value)
                            avg_per_game_against = self._safe_float(sheet.cell(row=row, column=9).value)
                            
                            ws_standings.append([
                                self.next_standing_id,
                                team_id,
                                season_id,
                                wins,
                                losses,
                                ties,
                                pins_for,
                                pins_against,
                                round(avg_per_game, 2) if avg_per_game > 0 else "",
                                round(avg_per_game_against, 2) if avg_per_game_against > 0 else "",
                                ""  # updated_at
                            ])
                            
                            self.next_standing_id += 1
            
            print(f"  {season_name}: Migrated standings")
    
    def migrate_champions(self):
        """Migrate season champions."""
        print("Migrating champions...")
        ws_champions = self.new_wb["season_champions"]
        
        # Clear example data
        for row in range(ws_champions.max_row, 1, -1):
            ws_champions.delete_rows(row)
        
        # Check if Champs sheet exists
        if "Champs" in self.old_wb.sheetnames:
            champs_sheet = self.old_wb["Champs"]
            
            for row in range(3, champs_sheet.max_row + 1):
                season_cell = champs_sheet.cell(row=row, column=2).value
                champ_cell = champs_sheet.cell(row=row, column=3).value
                
                if season_cell and champ_cell:
                    try:
                        season_num = int(season_cell)
                        season_name = f"Season {season_num}"
                        champ_team = str(champ_cell).strip()
                        
                        if season_name in self.seasons_map:
                            season_id = self.seasons_map[season_name]
                            # Find team_id
                            team_id = None
                            for (s_id, team_name), t_id in self.teams_map.items():
                                if s_id == season_id and champ_team.lower() in team_name.lower():
                                    team_id = t_id
                                    break
                            
                            if team_id:
                                ws_champions.append([
                                    self.next_standing_id,  # Reuse counter
                                    season_id,
                                    team_id,
                                    champ_team  # notes
                                ])
                                self.next_standing_id += 1
                    except (ValueError, TypeError):
                        continue
            
            print("  Migrated champions from Champs sheet")
    
    def migrate_all(self, detailed_seasons_start=10):
        """Run full migration."""
        print("=" * 60)
        print("Starting Migration")
        print("=" * 60)
        
        self.migrate_seasons()
        self.migrate_players()
        self.migrate_teams()
        self.migrate_team_players()
        self.migrate_games_and_aggregates(detailed_seasons_start)
        self.migrate_team_standings()
        self.migrate_champions()
        
        # Save the new workbook
        output_file = "Bowling_League_Database_Migrated.xlsx"
        self.new_wb.save(output_file)
        
        print("=" * 60)
        print(f"Migration complete! Saved to: {output_file}")
        print("=" * 60)
        print(f"Summary:")
        print(f"  Players: {len(self.players_map)}")
        print(f"  Teams: {len(self.teams_map)}")
        print(f"  Seasons: {len(self.seasons_map)}")
        print(f"  Games: {self.next_game_id - 1}")
        print(f"  Aggregates: {self.next_aggregate_id - 1}")


def main():
    """Main migration function."""
    import sys
    
    old_file = "Bowling- Friends League v4.xlsx"
    new_file = "Bowling_League_Database.xlsx"
    
    print("Excel Migration Tool")
    print("=" * 60)
    print(f"Source: {old_file}")
    print(f"Template: {new_file}")
    print()
    
    try:
        migrator = ExcelMigrator(old_file, new_file)
        
        # Determine which seasons have detailed game data
        # Default: Season 10+ have individual games, older are aggregates
        detailed_start = 10
        
        # Check if command line argument provided
        if len(sys.argv) > 1:
            try:
                detailed_start = int(sys.argv[1])
                print(f"Using season {detailed_start}+ for detailed games")
            except ValueError:
                print(f"Invalid season number, using default: {detailed_start}")
        else:
            # Interactive mode (optional)
            try:
                print("Which season number should start having detailed individual games?")
                print("(Older seasons will be stored as aggregates)")
                print("Default: 10 (press Enter to use default)")
                user_input = input("Season number: ").strip()
                if user_input:
                    detailed_start = int(user_input)
            except (EOFError, KeyboardInterrupt):
                # Non-interactive mode, use default
                print(f"Using default: Season {detailed_start}+ for detailed games")
        
        migrator.migrate_all(detailed_seasons_start=detailed_start)
        
    except FileNotFoundError as e:
        print(f"Error: File not found - {e}")
        print("Make sure both Excel files are in the current directory.")
    except Exception as e:
        print(f"Error during migration: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

