"""
Sheet handler module for reading/writing bowling league data.
Supports local Excel files with one row per week per player structure.
"""
import hashlib
from abc import ABC, abstractmethod
from typing import Any, Dict, Iterator, List, Optional
from openpyxl import load_workbook
from stats import compute
from utils import safe_float, safe_int


class SheetHandler(ABC):
    """Abstract base class for sheet handlers."""
    
    @abstractmethod
    def get_team_scores(
        self,
        team_name: Optional[str] = None,
        season: Optional[str] = None,
        week: Optional[int] = None,
        through_week: Optional[int] = None,
    ) -> Dict:
        """Get team scores. If team_name is None, return all teams.
        If through_week is set, aggregate only rows with Week <= through_week and exclude playoff-flagged rows.
        Do not pass both week and through_week."""
        pass
    
    @abstractmethod
    def get_player_scores(self, player_name: Optional[str] = None, season: Optional[str] = None, week: Optional[int] = None) -> Dict:
        """Get player scores. If player_name is None, return all players."""
        pass

    @abstractmethod
    def add_score(self, player_name: str, score: int, week: Optional[int] = None, season: Optional[str] = None) -> bool:
        """Add a score for a player. Returns True if successful."""
        pass
    
    @abstractmethod
    def get_seasons(self) -> List[str]:
        """Get list of available seasons."""
        pass


class ExcelHandler(SheetHandler):
    """Handler for local Excel files with one row per week per player structure."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self._load_workbook()
    
    def _load_workbook(self):
        """Load the Excel workbook."""
        self.workbook = load_workbook(self.file_path, data_only=True)
    
    def get_seasons(self) -> List[str]:
        """Get list of available season sheet names."""
        return [sheet for sheet in self.workbook.sheetnames if sheet.startswith('Season')]
    
    def _get_current_season(self) -> str:
        """Get the most recent season (highest number)."""
        seasons = self.get_seasons()
        if not seasons:
            return None
        # Sort by season number
        seasons.sort(key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 0, reverse=True)
        return seasons[0]
    
    def _get_season_number(self, season: Optional[str] = None) -> Optional[int]:
        """Get season number from season name."""
        if season is None:
            season = self._get_current_season()
        if season is None:
            return None
        try:
            return int(season.split()[-1])
        except (ValueError, IndexError):
            return None
    
    def _safe_float(self, value, default=0.0):
        return safe_float(value, default)

    def _safe_int(self, value, default=0):
        return safe_int(value, default)
    
    def _normalize(self, text: str) -> str:
        """Normalize text for comparison — lowercase and flatten curly quotes."""
        return text.lower().replace('\u2018', "'").replace('\u2019', "'").replace('\u201c', '"').replace('\u201d', '"')

    def _is_absent(self, value) -> bool:
        """Check if absent flag is set."""
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.upper() in ['Y', 'YES', 'TRUE', '1']
        return bool(value)
    
    def _is_substitute(self, value) -> bool:
        """Check if substitute flag is set."""
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.upper() in ['Y', 'YES', 'TRUE', '1']
        return bool(value)

    def _is_playoff_row(self, value) -> bool:
        """True if v5 'Playoffs?' column (col 12) marks this player-week as playoffs."""
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().upper() in ['Y', 'YES', 'TRUE', '1']
        return bool(value)

    def _optional_score(self, value) -> Optional[float]:
        if value is None or value == "":
            return None
        score = self._safe_float(value)
        return score if score > 0 else None

    def _row_fingerprint(
        self,
        sheet_key: str,
        week: int,
        team: str,
        player: str,
        games: tuple,
        week_average: Optional[float],
        absent: bool,
        substitute: bool,
        playoffs: bool,
        opponent: Optional[str],
    ) -> str:
        parts = (
            f"{sheet_key}|{week}|{team}|{player}|"
            f"{games[0]}|{games[1]}|{games[2]}|{games[3]}|{games[4]}|"
            f"{week_average}|{absent}|{substitute}|{playoffs}|{opponent or ''}"
        )
        return hashlib.sha256(parts.encode()).hexdigest()[:64]

    def iter_player_week_rows(
        self, season_filter: Optional[str] = None
    ) -> Iterator[Dict[str, Any]]:
        """Yield one normalized dict per v5 sheet row (all seasons unless season_filter is set)."""
        sheet_keys = [
            s
            for s in self.get_seasons()
            if s.startswith("Season")
        ]
        sheet_keys.sort(key=lambda x: self._safe_int(x.split()[-1], 0))

        if season_filter is not None:
            num = self._safe_int(season_filter.split()[-1], -1)
            if num < 0:
                num = self._safe_int(season_filter, -1)
            if num < 0:
                return
            sheet_keys = [s for s in sheet_keys if self._safe_int(s.split()[-1], -1) == num]

        for sheet_key in sheet_keys:
            season_num = self._get_season_number(sheet_key)
            if season_num is None or sheet_key not in self.workbook.sheetnames:
                continue

            sheet = self.workbook[sheet_key]
            for row in range(2, sheet.max_row + 1):
                row_team = sheet.cell(row=row, column=2).value
                row_player = sheet.cell(row=row, column=3).value
                row_season = sheet.cell(row=row, column=4).value
                row_week = sheet.cell(row=row, column=5).value

                if self._safe_int(row_season, -1) != season_num:
                    continue

                week = self._safe_int(row_week, 0)
                if week <= 0:
                    continue

                if not row_player or not isinstance(row_player, str):
                    continue
                player = row_player.strip()
                if not player or player.lower() in ("player", "team"):
                    continue

                team = str(row_team).strip() if row_team else ""
                if not team or team.lower() == "team":
                    continue

                games = tuple(
                    self._optional_score(sheet.cell(row=row, column=col).value)
                    for col in range(6, 11)
                )
                avg_cell = sheet.cell(row=row, column=11).value
                week_average = None
                if avg_cell is not None and avg_cell != "":
                    week_average = self._safe_float(avg_cell)
                elif any(g is not None for g in games):
                    played = [g for g in games if g is not None]
                    week_average = sum(played) / len(played) if played else None

                absent = self._is_absent(sheet.cell(row=row, column=14).value)
                substitute = self._is_substitute(sheet.cell(row=row, column=15).value)
                playoffs = self._is_playoff_row(sheet.cell(row=row, column=12).value)
                opponent_cell = sheet.cell(row=row, column=16).value
                opponent = str(opponent_cell).strip() if opponent_cell else None
                g5_cell = sheet.cell(row=row, column=13).value
                game5_winner = str(g5_cell).strip() if g5_cell else None

                yield {
                    "sheet_key": sheet_key,
                    "season_number": season_num,
                    "season_label": sheet_key,
                    "team": team,
                    "player_display_name": player,
                    "week": week,
                    "game1": games[0],
                    "game2": games[1],
                    "game3": games[2],
                    "game4": games[3],
                    "game5": games[4],
                    "week_average": week_average,
                    "absent": absent,
                    "substitute": substitute,
                    "playoffs": playoffs,
                    "opponent": opponent,
                    "game5_winner": game5_winner,
                    "source_row_fingerprint": self._row_fingerprint(
                        sheet_key,
                        week,
                        team,
                        player,
                        games,
                        week_average,
                        absent,
                        substitute,
                        playoffs,
                        opponent,
                    ),
                }

    def _facts_for_season(self, season: Optional[str] = None) -> List[Dict[str, Any]]:
        return list(self.iter_player_week_rows(season))

    def _all_facts(self) -> List[Dict[str, Any]]:
        return list(self.iter_player_week_rows())

    def get_team_scores(
        self,
        team_name: Optional[str] = None,
        season: Optional[str] = None,
        week: Optional[int] = None,
        through_week: Optional[int] = None,
    ) -> Dict:
        """Get team scores from Excel. Team average is average of individual player averages.
        Total pins includes absences but excludes substitutes.
        Calculates wins/losses/ties from weekly matchups.
        If week is specified, returns individual games for that week.
        If through_week is set (and week is None), stats include only weeks 1..through_week (excluding playoff rows).
        If season is None, uses the latest/current season."""
        if season is None:
            season = self._get_current_season()
        season_num = self._get_season_number(season)
        if season_num is None:
            return {"error": f"Season '{season}' not found"}
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return {"error": f"Season '{season}' not found"}
        return compute.get_team_scores(
            self._facts_for_season(season),
            team_name,
            season,
            week,
            through_week,
            season_num=season_num,
        )

    def get_team_weekly_summary(self, team_name: str, season: Optional[str] = None) -> Dict:
        """Get weekly breakdown for a team showing opponent, record, and totals per week."""
        season_num = self._get_season_number(season)
        if season_num is None:
            return {"error": f"Season '{season}' not found"}
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return {"error": f"Season '{season}' not found"}
        return compute.get_team_weekly_summary(
            self._facts_for_season(season), team_name, season, season_num=season_num,
        )

    def get_league_stats(self, season: Optional[str] = None) -> Dict:
        """Get league statistics including top players, best weeks, best team totals, and best games."""
        season_num = self._get_season_number(season)
        if season_num is None:
            return {"error": f"Season '{season}' not found"}
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return {"error": f"Season '{season}' not found"}
        return compute.get_league_stats(
            self._facts_for_season(season), season, season_num=season_num,
        )

    def get_all_time_stats(self) -> Dict:
        """Aggregate league stats across all seasons for all-time leaders."""
        return compute.get_all_time_stats(self._all_facts())

    def get_player_scores(self, player_name: Optional[str] = None, season: Optional[str] = None, week: Optional[int] = None) -> Dict:
        """Get player scores from Excel. Excludes absent weeks from average calculation."""
        if season is None:
            season = self._get_current_season()
        season_num = self._get_season_number(season)
        if season_num is None:
            return {"error": f"Season '{season}' not found"}
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return {"error": f"Season '{season}' not found"}
        return compute.get_player_scores(
            self._facts_for_season(season), player_name, season, week, season_num=season_num,
        )

    def add_score(self, player_name: str, score: int, week: Optional[int] = None, season: Optional[str] = None) -> bool:
        """Add a score for a player. Note: This modifies the Excel file.

        LIMITATION: The workbook is loaded with data_only=True, which caches formula results
        as static values. Saving after a write will permanently strip any Excel formulas in the
        file (e.g. the Average column). Do not use this method if you rely on Excel formulas
        being preserved. Manual edits in Excel are the recommended way to update scores.
        """
        season_num = self._get_season_number(season)
        if season_num is None:
            return False
        
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return False
        
        sheet = self.workbook[season_sheet]
        
        # Find the player's row for the specified week (or latest week if not specified)
        target_row = None
        
        if week is None:
            # Find the latest week for this player
            max_week = 0
            for row in range(2, sheet.max_row + 1):
                row_player = sheet.cell(row=row, column=3).value
                row_season = sheet.cell(row=row, column=4).value
                row_week = sheet.cell(row=row, column=5).value
                
                if (row_player and isinstance(row_player, str) and 
                    player_name.lower() in row_player.lower() and 
                    row_season == season_num):
                    week_num = self._safe_int(row_week, 0)
                    if week_num > max_week:
                        max_week = week_num
                        target_row = row
            
            if target_row:
                # Find first empty game column
                for col in range(6, 11):  # Game 1-5 columns
                    game_value = sheet.cell(row=target_row, column=col).value
                    if game_value is None or game_value == "":
                        sheet.cell(row=target_row, column=col, value=score)
                        self.workbook.save(self.file_path)
                        return True
        else:
            # Find specific week for this player
            for row in range(2, sheet.max_row + 1):
                row_player = sheet.cell(row=row, column=3).value
                row_season = sheet.cell(row=row, column=4).value
                row_week = sheet.cell(row=row, column=5).value
                
                if (row_player and isinstance(row_player, str) and 
                    player_name.lower() in row_player.lower() and 
                    row_season == season_num and 
                    self._safe_int(row_week, 0) == week):
                    
                    # Find first empty game column
                    for col in range(6, 11):  # Game 1-5 columns
                        game_value = sheet.cell(row=row, column=col).value
                        if game_value is None or game_value == "":
                            sheet.cell(row=row, column=col, value=score)
                            self.workbook.save(self.file_path)
                            return True
                    break
        
        return False

    def get_latest_week(self, season: Optional[str] = None) -> int:
        """Return the highest week number that has any data in the given season."""
        if season is None:
            season = self._get_current_season()
        season_num = self._get_season_number(season)
        if season_num is None:
            return 1
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return 1
        return compute.get_latest_week(
            self._facts_for_season(season), season, season_num=season_num,
        )

    def list_weeks_for_season(self, season: Optional[str] = None) -> List[int]:
        """Sorted distinct week numbers that have at least one row for this season."""
        if season is None:
            season = self._get_current_season()
        season_num = self._get_season_number(season)
        if season_num is None:
            return []
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return []
        return compute.list_weeks_for_season(
            self._facts_for_season(season), season, season_num=season_num,
        )

    def list_playoff_weeks_for_season(self, season: Optional[str] = None) -> List[int]:
        """Week numbers to show as playoff weeks."""
        return compute.list_playoff_weeks_for_season(
            self._all_facts(), season, seasons=self.get_seasons(),
        )

    def get_week_summary(self, week: int, season: Optional[str] = None) -> dict:
        """Return all player and league stats for a specific week."""
        if season is None:
            season = self._get_current_season()
        season_num = self._get_season_number(season)
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return {"error": f"Season '{season}' not found"}
        return compute.get_week_summary(
            self._facts_for_season(season), week, season, season_num=season_num,
        )

    def get_week_matchups(self, week: int, season: Optional[str] = None) -> dict:
        """Return team matchup results for a specific week."""
        if season is None:
            season = self._get_current_season()
        season_num = self._get_season_number(season)
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return {"error": f"Season '{season}' not found"}
        return compute.get_week_matchups(
            self._facts_for_season(season), week, season, season_num=season_num,
        )

    def find_player_names(self, search: str, season: Optional[str] = None) -> List[str]:
        """Return all unique player names that match the search term."""
        if season is None:
            season = self._get_current_season()
        season_num = self._get_season_number(season)
        if season_num is None:
            return []
        season_sheet = f"Season {season_num}"
        if season_sheet not in self.workbook.sheetnames:
            return []
        return compute.find_player_names(
            self._facts_for_season(season), search, season, season_num=season_num,
        )


# ---------------------------------------------------------------------------
# Factory (sync_db / import only — web app reads PostgreSQL)
# ---------------------------------------------------------------------------

def get_sheet_handler(handler_type: str = "excel", **kwargs) -> SheetHandler:
    """Local Excel workbook for sync_db. The web app does not call this."""
    if handler_type.lower() != "excel":
        raise ValueError(f"Only excel handler is supported (got {handler_type!r})")
    if "file_path" not in kwargs:
        raise ValueError("file_path is required for Excel handler")
    return ExcelHandler(kwargs["file_path"])
