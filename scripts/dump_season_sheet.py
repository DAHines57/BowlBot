"""Dump one season sheet from Excel. Usage: python scripts/dump_season_sheet.py 3"""
from __future__ import annotations

import os
import sys
from collections import defaultdict

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
PATH = os.environ.get("EXCEL_FILE_PATH", "Bowling-Friends League v5.xlsx")


def main() -> None:
    sn = int(sys.argv[1]) if len(sys.argv) > 1 else 4
    sheet = f"Season {sn}"
    wb = load_workbook(PATH, data_only=True)
    ws = wb[sheet]
    print(f"=== {sheet} ===\n")
    by_week: dict[int, list] = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 4).value != sn:
            continue
        wk = int(ws.cell(row, 5).value or 0)
        if wk <= 0:
            continue
        opp = ws.cell(row, 16).value
        po = ws.cell(row, 12).value
        by_week[wk].append((str(ws.cell(row, 2).value).strip(), opp, po))
    for wk in sorted(by_week):
        rows = by_week[wk]
        po = any(r[2] for r in rows)
        opps = {r[1] for r in rows if r[1]}
        print(f"Week {wk} playoffs={po} distinct opponents={len(opps)}")
        if wk >= max(by_week) - 2:
            pairs = set()
            for team, opp, _ in rows:
                if opp:
                    pairs.add(tuple(sorted((team, str(opp).strip()))))
            for p in sorted(pairs):
                print(f"  {p[0]} vs {p[1]}")
    wb.close()


if __name__ == "__main__":
    main()
