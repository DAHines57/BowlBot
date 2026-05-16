"""Dump Season 4 playoff structure from the v5 Excel sheet."""
from __future__ import annotations

import os
from collections import defaultdict

from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
PATH = os.environ.get("EXCEL_FILE_PATH", "Bowling-Friends League v5.xlsx")
SHEET = "Season 4"


def playoff_flag(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("Y", "YES", "TRUE", "1")


def main() -> None:
    wb = load_workbook(PATH, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"Sheet {SHEET!r} not found. Sheets:", wb.sheetnames[:15])
        return
    ws = wb[SHEET]
    print(f"=== {SHEET} (rows 1-{ws.max_row}, cols 1-{ws.max_column}) ===\n")
  # Header
    headers = [ws.cell(1, c).value for c in range(1, min(ws.max_column + 1, 20))]
    print("Row 1 headers:", headers)

    by_week: dict[int, list[dict]] = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        season = ws.cell(row, 4).value
        if season != 4 and str(season) != "4":
            continue
        week = ws.cell(row, 5).value
        try:
            wk = int(week)
        except (TypeError, ValueError):
            continue
        team = ws.cell(row, 2).value
        player = ws.cell(row, 3).value
        if not team or not player:
            continue
        playoffs = playoff_flag(ws.cell(row, 12).value)
        g5w = ws.cell(row, 13).value
        opponent = ws.cell(row, 16).value
        games = [ws.cell(row, c).value for c in range(6, 11)]
        by_week[wk].append(
            {
                "team": str(team).strip(),
                "player": str(player).strip(),
                "playoffs": playoffs,
                "opponent": str(opponent).strip() if opponent else None,
                "g5_winner": str(g5w).strip() if g5w else None,
                "games": games,
            }
        )

    for wk in sorted(by_week):
        rows = by_week[wk]
        po = any(r["playoffs"] for r in rows)
        print(f"\n--- Week {wk} (playoffs={po}, rows={len(rows)}) ---")
        teams = sorted({r["team"] for r in rows})
        print("Teams:", ", ".join(teams))
        pairs: dict[tuple[str, str], int] = {}
        for r in rows:
            if not r["opponent"]:
                continue
            key = tuple(sorted((r["team"], r["opponent"])))
            pairs[key] = pairs.get(key, 0) + 1
        print("Matchups (from opponent column):")
        for (a, b), n in sorted(pairs.items()):
            print(f"  {a} vs {b}  ({n} player rows)")

    # Scan for non-tall matchup blocks (cols 17+)
    print("\n--- Non-empty cells in cols 17-30, rows 1-80 (sample) ---")
    for row in range(1, min(81, ws.max_row + 1)):
        for col in range(17, min(31, ws.max_column + 1)):
            v = ws.cell(row, col).value
            if v is not None and str(v).strip():
                print(f"  R{row}C{col}: {v!r}")
    wb.close()


if __name__ == "__main__":
    main()
