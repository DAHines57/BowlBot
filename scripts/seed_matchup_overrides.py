"""
Load matchup_overrides from v4 matchup columns (Season 4+).

Requires DATABASE_URL and Bowling- Friends League v4.xlsx in the project root.

  python scripts/seed_matchup_overrides.py
  python scripts/seed_matchup_overrides.py --dry-run
  python scripts/seed_matchup_overrides.py --season 9
"""
from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete

_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

from db.excluded_seasons import is_season_excluded
from db.models import MatchupOverride

# v4 workbook tabs that include a matchup W/L block (Seasons 4–13).
V4_SHEETS_WITH_MATCHUP_BLOCK: frozenset[int] = frozenset(range(4, 14))
from db.session import get_session
from db.sync import _get_or_create_season
from migrate import V4_FILE, v4_matchup_overrides_for_sheet
from stats.facts import canonical_team_name, resolve_opponent_on_roster

logger = logging.getLogger(__name__)


def _normalize_overrides(raw: list[dict], season_num: int) -> list[dict]:
    roster = sorted(
        {
            canonical_team_name(r["team"], season_num=season_num)
            for r in raw
        }
    )
    out: list[dict] = []
    for r in raw:
        team = canonical_team_name(r["team"], season_num=season_num)
        opp_raw = str(r["opponent"]).strip()
        opponent = (
            resolve_opponent_on_roster(opp_raw, roster, season_num=season_num) or opp_raw
        )
        out.append(
            {
                "week": int(r["week"]),
                "team": team,
                "opponent": opponent,
                "wins": int(r["wins"]),
                "losses": int(r["losses"]),
                "ties": int(r["ties"]),
                "playoffs": bool(r["playoffs"]),
            }
        )
    return out


def seed_from_v4(
    *,
    dry_run: bool = False,
    season_filter: int | None = None,
    v4_path: Path | None = None,
) -> dict:
    path = v4_path or (_ROOT / V4_FILE)
    if not path.is_file():
        raise FileNotFoundError(f"v4 workbook not found: {path}")

    wb = load_workbook(path, data_only=True)
    season_sheets = [
        name
        for name in wb.sheetnames
        if name.strip().lower().startswith("season")
        and re.search(r"(\d+)", name)
    ]

    total_rows = 0
    seasons_touched = 0

    with get_session() as session:
        for sheet_name in sorted(
            season_sheets,
            key=lambda n: int(re.search(r"(\d+)", n).group(1)),
        ):
            season_num = int(re.search(r"(\d+)", sheet_name).group(1))
            if season_filter is not None and season_num != season_filter:
                continue
            if season_num not in V4_SHEETS_WITH_MATCHUP_BLOCK:
                logger.info("Season %s: no v4 matchup section — skip", season_num)
                continue
            if is_season_excluded(season_num):
                logger.info("Season %s: excluded — skip", season_num)
                continue

            raw = v4_matchup_overrides_for_sheet(wb[sheet_name])
            if not raw:
                logger.warning("Season %s: matchup section empty — skip", season_num)
                continue

            rows = _normalize_overrides(raw, season_num)
            total_rows += len(rows)
            seasons_touched += 1
            logger.info("Season %s: %s override rows", season_num, len(rows))

            if dry_run:
                continue

            season = _get_or_create_season(session, sheet_name, season_num)
            session.execute(
                delete(MatchupOverride).where(MatchupOverride.season_id == season.id)
            )
            for r in rows:
                session.add(
                    MatchupOverride(
                        season_id=season.id,
                        week=r["week"],
                        team=r["team"],
                        opponent=r["opponent"],
                        wins=r["wins"],
                        losses=r["losses"],
                        ties=r["ties"],
                        playoffs=r["playoffs"],
                    )
                )

        if not dry_run:
            session.commit()

    return {
        "dry_run": dry_run,
        "rows": total_rows,
        "seasons": seasons_touched,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Seed matchup_overrides from v4 matchup columns"
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--season", type=int, metavar="N")
    parser.add_argument("-v", "--verbose", action="store_true")
    parser.add_argument(
        "--v4-file",
        type=Path,
        default=None,
        help=f"Path to v4 workbook (default: {_ROOT / V4_FILE})",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )

    try:
        result = seed_from_v4(
            dry_run=args.dry_run,
            season_filter=args.season,
            v4_path=args.v4_file,
        )
    except Exception as exc:
        logger.error("%s", exc)
        return 1

    mode = "dry-run" if result["dry_run"] else "seed"
    print(
        f"{mode}: {result['rows']} rows across {result['seasons']} season(s)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
