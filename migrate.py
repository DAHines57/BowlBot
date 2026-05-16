"""
migrate.py — One-time migration script from v4 to v5 Excel format.

v4 format (wide):  one row per player, weekly averages in columns
v5 format (tall):  one row per player per week, with game scores and opponent.
    Optional column **Game 5 winner** (after Playoffs?) holds the series winner when
    the v4 matchup row shows a 5-game series (W+L[+T]=5) but pin scores for game 5 are unknown.

Since v4 only stores weekly averages (not individual games), each week's
average is repeated as Game 1–4 so season averages calculate correctly.
Game 5 is left blank (v4 has no 5-game data).

Output: Bowling-Friends League migrated.xlsx
  — one sheet per season, ready to copy into v5 / Google Sheets as needed.

Google Sheets (live app source)
  The app reads any worksheet whose name starts with ``Season`` (e.g. ``Season 13``).
  This script does not push to Google Drive. To add **Season 13** there:

  1. Put ``Season 13`` on the v4 workbook (``V4_FILE`` below), OR set ``ONLY_SEASON_NUMBER = 13``
     if that tab is the only new one you need.
  2. Run: ``python migrate.py``
  3. Open ``Bowling-Friends League migrated.xlsx``, copy the whole ``Season 13`` sheet
     into your Google Sheet (same tab name), or use Sheets import.

  Optional: set ``ONLY_SEASON_NUMBER = 13`` below to migrate only that season.
"""

import re
from colorsys import hls_to_rgb, rgb_to_hls
from typing import List, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.xml.functions import QName, fromstring

# No background — Game 5 winner cells stay clear when empty or unmatched
_NO_FILL = PatternFill(fill_type='none')

V4_FILE  = "Bowling- Friends League v4.xlsx"
OUTPUT_FILE = "Bowling-Friends League migrated.xlsx"

# Seasons to omit from migration (e.g. if you maintain a tab by hand). Empty = migrate all.
SKIP_SEASONS = set()

# Seasons where weekly values are 4-game pin totals instead of averages.
# The script will divide by 4 before writing game scores.
TOTAL_SEASONS = {3}

# If set, migrate only this season (e.g. 13). None = migrate every season except SKIP_SEASONS.
ONLY_SEASON_NUMBER: Optional[int] = None

V5_HEADER = [
    "Index", "Team", "Player", "Season", "Week",
    "Game 1", "Game 2", "Game 3", "Game 4", "Game 5",
    "Average", "Playoffs?", "Game 5 winner", "Absent?", "Substitute?", "Opponent"
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_week_label(label):
    """
    Parse a week column header into (week_number, is_playoff).
    e.g. 'Week 3' -> (3, False), 'Playoff Week 2' -> (2, True)
    Returns (None, False) if the label doesn't match.
    """
    if not label:
        return None, False
    s = str(label).strip()
    if re.search(r'playoff', s, re.IGNORECASE):
        m = re.search(r'(\d+)', s)
        return (int(m.group(1)), True) if m else (None, False)
    m = re.search(r'week\s+(\d+)', s, re.IGNORECASE)
    return (int(m.group(1)), False) if m else (None, False)


def find_vs_offset(ws, row, start_col, end_col):
    """Return the column offset (relative to start_col) where 'Vs' appears."""
    for c in range(start_col, end_col + 1):
        if ws.cell(row, c).value == 'Vs':
            return c - start_col
    return None


def parse_matchups(ws, matchup_start_col, max_regular_week):
    """
    Parse the matchup section and return a dict:
        (team_name_lower, absolute_week_number) -> opponent_name

    Handles both formats:
      - With Ties:    TeamA, W, L, T, Pins, Vs, TeamB, ...
      - Without Ties: TeamA, W, L, Pins, Vs, TeamB, ...
    """
    matchups = {}
    current_week = None

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row, matchup_start_col).value
        if cell_val is None:
            continue
        cell_str = str(cell_val).strip()

        # Week header line?
        wn, is_playoff = parse_week_label(cell_str)
        if wn is not None:
            current_week = max_regular_week + wn if is_playoff else wn
            continue

        if current_week is None:
            continue

        # Try to find 'Vs' somewhere in this row
        vs_offset = find_vs_offset(ws, row, matchup_start_col, matchup_start_col + 8)
        if vs_offset is None:
            continue

        team_a = ws.cell(row, matchup_start_col).value
        team_b = ws.cell(row, matchup_start_col + vs_offset + 1).value

        if not (team_a and isinstance(team_a, str) and
                team_b and isinstance(team_b, str)):
            continue

        # Skip bracket/placement label rows (e.g. "Winners Bracket")
        # Real team rows have a numeric value in column 2 (Wins)
        wins_a = ws.cell(row, matchup_start_col + 1).value
        if not isinstance(wins_a, (int, float)):
            continue

        team_a = team_a.strip()
        team_b = team_b.strip()
        matchups[(team_a.lower(), current_week)] = team_b
        matchups[(team_b.lower(), current_week)] = team_a

    return matchups


def _int_cell(ws, row, col):
    v = ws.cell(row, col).value
    if isinstance(v, (int, float)):
        return int(v)
    return None


def parse_game5_winners(ws, matchup_start_col, max_regular_week):
    """
    From v4 matchup rows: if each team's wins+losses (+ ties when present) totals 5,
    the series went five games. Record the deciding-game (series) winner's team name
    for both teams for that absolute week.

    Returns:
        dict: (team_name_lower, absolute_week) -> winner display name, or absent key if no G5.
    """
    out = {}
    current_week = None

    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row, matchup_start_col).value
        if cell_val is None:
            continue
        cell_str = str(cell_val).strip()

        wn, is_playoff = parse_week_label(cell_str)
        if wn is not None:
            current_week = max_regular_week + wn if is_playoff else wn
            continue

        if current_week is None:
            continue

        vs_offset = find_vs_offset(ws, row, matchup_start_col, matchup_start_col + 12)
        if vs_offset is None:
            continue

        team_a = ws.cell(row, matchup_start_col).value
        team_b = ws.cell(row, matchup_start_col + vs_offset + 1).value

        if not (team_a and isinstance(team_a, str) and
                team_b and isinstance(team_b, str)):
            continue

        wins_a = ws.cell(row, matchup_start_col + 1).value
        if not isinstance(wins_a, (int, float)):
            continue

        w_a = _int_cell(ws, row, matchup_start_col + 1)
        l_a = _int_cell(ws, row, matchup_start_col + 2)
        w_b = _int_cell(ws, row, matchup_start_col + vs_offset + 2)
        l_b = _int_cell(ws, row, matchup_start_col + vs_offset + 3)
        if None in (w_a, l_a, w_b, l_b):
            continue

        # With ties: Team, W, L, T, Pins, Vs — 'Vs' is one column farther than without ties
        if vs_offset == 5:
            t_a = _int_cell(ws, row, matchup_start_col + 3) or 0
            t_b = _int_cell(ws, row, matchup_start_col + vs_offset + 4) or 0
            games_a = w_a + l_a + t_a
            games_b = w_b + l_b + t_b
        else:
            games_a = w_a + l_a
            games_b = w_b + l_b

        if games_a != 5 or games_b != 5:
            continue
        if w_a == w_b:
            continue

        team_a = team_a.strip()
        team_b = team_b.strip()
        winner = team_a if w_a > w_b else team_b
        out[(team_a.lower(), current_week)] = winner
        out[(team_b.lower(), current_week)] = winner

    return out


# ---------------------------------------------------------------------------
# Per-season migration
# ---------------------------------------------------------------------------

def parse_formula_games(formula):
    """
    Extract individual game scores from a formula like =(171+203+126+166)/4.
    Returns a list of floats on success, None if the pattern doesn't match.
    """
    if not formula or not isinstance(formula, str):
        return None
    m = re.match(r'=\(([\d.]+(?:\+[\d.]+)+)\)/\d+', formula.strip())
    if m:
        return [float(x) for x in m.group(1).split('+')]
    return None


def _cell_is_highlighted(ws_styles, row, col):
    """Return True if the cell has any fill (indicating an absent week)."""
    cell = ws_styles.cell(row, col)
    ft = cell.fill.fill_type
    return ft is not None and ft != 'none'


# --- Excel fill → aRGB (theme / indexed / rgb) --------------------------------

_RGBMAX = 0xFF
_HLSMAX = 240


def _rgb_to_ms_hls(hex6: str):
    """6-digit RRGGBB (no alpha) → Excel HLS triple on 0..240 scale."""
    hex6 = hex6[-6:]
    r = int(hex6[0:2], 16) / _RGBMAX
    g = int(hex6[2:4], 16) / _RGBMAX
    b = int(hex6[4:6], 16) / _RGBMAX
    h, l, s = rgb_to_hls(r, g, b)
    return (
        int(round(h * _HLSMAX)),
        int(round(l * _HLSMAX)),
        int(round(s * _HLSMAX)),
    )


def _ms_hls_to_hex6(h: int, l: int, s: int) -> str:
    r01, g01, b01 = hls_to_rgb(h / _HLSMAX, l / _HLSMAX, s / _HLSMAX)
    return (
        f"{int(round(r01 * _RGBMAX)):02X}"
        f"{int(round(g01 * _RGBMAX)):02X}"
        f"{int(round(b01 * _RGBMAX)):02X}"
    )


def _tint_luminance(tint: float, lum: int) -> int:
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    return int(round(lum * (1.0 - tint) + (_HLSMAX - _HLSMAX * (1.0 - tint))))


def _theme_palette_from_workbook(wb) -> Optional[List[str]]:
    """Parse theme accent/base colors as 6-char RRGGBB from the workbook."""
    raw = getattr(wb, "loaded_theme", None)
    if not raw:
        return None
    try:
        xlmns = "http://schemas.openxmlformats.org/drawingml/2006/main"
        root = fromstring(raw)
        theme_el = root.find(QName(xlmns, "themeElements").text)
        schemes = theme_el.findall(QName(xlmns, "clrScheme").text)
        first = schemes[0]
        out = []
        for name in (
            "lt1", "dk1", "lt2", "dk2",
            "accent1", "accent2", "accent3", "accent4", "accent5", "accent6",
        ):
            el = first.find(QName(xlmns, name).text)
            for child in list(el):
                v = child.attrib.get("val", "")
                if "window" in v:
                    out.append(child.attrib["lastClr"])
                else:
                    out.append(child.attrib["val"])
        return out
    except Exception:
        return None


def _theme_and_tint_to_argb(hex6: str, tint: float) -> str:
    h, l, s = _rgb_to_ms_hls(hex6)
    l2 = _tint_luminance(tint, l)
    body = _ms_hls_to_hex6(h, l2, s)
    return "FF" + body


def _pattern_fill_fg_argb(fill, theme_palette: Optional[List[str]]) -> Optional[str]:
    """
    Return 8-char aRGB (e.g. FFFF7C80) for PatternFill.fgColor, or None.
    Handles rgb, indexed, and theme (+ tint) like the v4 league workbook uses.
    """
    if not fill or fill.fill_type is None or fill.fill_type == 'none':
        return None
    fg = fill.fgColor
    t = fg.type

    if t == 'rgb' and fg.rgb:
        raw = str(fg.rgb).strip().upper()
        if raw in ('00000000', '00', '0'):
            return None
        if len(raw) == 8:
            return raw
        if len(raw) == 6:
            return "FF" + raw
        return raw

    if t == 'indexed' and fg.indexed is not None:
        idx = int(fg.indexed)
        if idx < 0 or idx >= len(COLOR_INDEX):
            return None
        entry = COLOR_INDEX[idx]
        return entry if len(entry) == 8 else "FF" + entry[-6:].upper()

    if t == 'theme' and fg.theme is not None and theme_palette:
        ti = int(fg.theme)
        if 0 <= ti < len(theme_palette):
            tint = float(fg.tint or 0.0)
            return _theme_and_tint_to_argb(theme_palette[ti], tint)

    return None


def _get_team_color(ws_styles, row, team_col, theme_palette: Optional[List[str]]):
    """
    Return the ARGB color string for a team row, or None if no usable fill.
    """
    cell = ws_styles.cell(row, team_col)
    return _pattern_fill_fg_argb(cell.fill, theme_palette)


def build_team_colors(ws_styles, team_col, player_col, theme_palette: Optional[List[str]]):
    """Return a dict of {team_name_lower: rgb_color} from the style sheet."""
    colors = {}
    for r in range(3, ws_styles.max_row + 1):
        team = ws_styles.cell(r, team_col).value
        if not team or not isinstance(team, str):
            continue
        color = _get_team_color(ws_styles, r, team_col, theme_palette)
        if color:
            colors[team.strip().lower()] = color
    return colors


def _team_color_for_name(team_colors: dict, name: str) -> Optional[str]:
    """Resolve aRGB fill for a roster team name; exact key then substring match."""
    key = name.strip().lower()
    if not key:
        return None
    rgb = team_colors.get(key)
    if rgb:
        return rgb
    for tk, rgb in team_colors.items():
        if tk in key or key in tk:
            return rgb
    return None


def migrate_season(ws, ws_styles, season_num, wb_book):
    """
    Read one v4 season sheet and return a list of v5 rows (as plain lists).
    """
    header = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

    # Locate the player section by finding 'Player' column
    player_col = None
    for i, val in enumerate(header):
        if val == 'Player':
            player_col = i + 1  # 1-based
            break

    if player_col is None:
        print(f"  WARNING: 'Player' column not found — skipping Season {season_num}")
        return []

    team_col = player_col - 1  # 'Team' is immediately left of 'Player'

    # Collect week columns and locate matchup start
    week_cols = {}       # column_index -> (week_number, is_playoff)
    avg_col = None
    matchup_start_col = None

    for i, val in enumerate(header):
        col = i + 1
        if col <= player_col:
            continue  # only look right of Player
        if isinstance(val, str):
            if 'Matchup' in val and matchup_start_col is None:
                # Must be checked before parse_week_label — "Week 1 Matchups"
                # contains "Week 1" and would be misidentified as a week column.
                matchup_start_col = col
            elif val == 'Average' and avg_col is None:
                avg_col = col
            else:
                wn, is_p = parse_week_label(val)
                if wn is not None:
                    week_cols[col] = (wn, is_p)

    if not week_cols:
        print(f"  WARNING: No week columns found — skipping Season {season_num}")
        return []

    # Compute max regular week for playoff offset
    max_regular_week = max(
        wn for wn, is_p in week_cols.values() if not is_p
    )

    # Build absolute week map: col -> absolute_week_num
    col_to_abs_week = {}
    col_to_is_playoff = {}
    for col, (wn, is_p) in week_cols.items():
        abs_week = max_regular_week + wn if is_p else wn
        col_to_abs_week[col] = abs_week
        col_to_is_playoff[col] = is_p

    # Parse opponent lookup
    matchups = {}
    game5_winners = {}
    if matchup_start_col:
        matchups = parse_matchups(ws, matchup_start_col, max_regular_week)
        game5_winners = parse_game5_winners(ws, matchup_start_col, max_regular_week)

    theme_palette = _theme_palette_from_workbook(wb_book)

    # Build team color lookup (team_name_lower -> rgb)
    team_colors = build_team_colors(ws_styles, team_col, player_col, theme_palette)

    # Iterate player rows
    rows_out = []
    index = 1

    for row in range(3, ws.max_row + 1):
        team = ws.cell(row, team_col).value
        player = ws.cell(row, player_col).value

        if not player or not isinstance(player, str):
            continue
        if not team or not isinstance(team, str):
            continue

        team = team.strip()
        player = player.strip()

        # Skip header/label rows that accidentally pass the string check
        if player.lower() in ('player', 'team', 'wins', 'losses', 'average'):
            continue

        for col in sorted(col_to_abs_week.keys()):
            abs_week = col_to_abs_week[col]
            is_playoff = col_to_is_playoff[col]
            weekly_avg = ws.cell(row, col).value

            # Non-numeric value in a week cell means stray data — skip it
            if weekly_avg is not None and not isinstance(weekly_avg, (int, float)):
                continue

            # A highlighted cell = player was absent (their average was substituted).
            # Fall back to None/0 check for seasons without consistent highlighting.
            is_absent = (
                _cell_is_highlighted(ws_styles, row, col)
                or weekly_avg is None
                or weekly_avg == 0
            )
            opponent = matchups.get((team.lower(), abs_week), "")
            game5_winner = game5_winners.get((team.lower(), abs_week), "")

            # Extract game scores for every row (absent or not).
            # ws_styles holds the raw formula string (loaded without data_only).
            formula = ws_styles.cell(row, col).value
            games = parse_formula_games(formula)

            if games and len(games) >= 4:
                g1 = round(games[0])
                g2 = round(games[1])
                g3 = round(games[2])
                g4 = round(games[3])
                g5 = round(games[4]) if len(games) >= 5 else None
                avg = round(sum(games) / len(games), 2)
            elif weekly_avg is not None:
                raw = float(weekly_avg)
                avg = round(raw / 4, 2) if season_num in TOTAL_SEASONS else round(raw, 2)
                g1 = g2 = g3 = g4 = avg
                g5 = None
            else:
                g1 = g2 = g3 = g4 = g5 = avg = None

            v5_row = [
                index, team, player, season_num, abs_week,
                g1, g2, g3, g4, g5,
                avg,
                "Y" if is_playoff else "N",
                game5_winner,
                "Y" if is_absent else "N", "N",
                opponent
            ]

            team_color = _get_team_color(ws_styles, row, team_col, theme_palette)
            rows_out.append((v5_row, team_color, opponent, team_colors))
            index += 1

    return rows_out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"Reading {V4_FILE} ...")
    # Load twice: data_only for cached numeric values, normal for cell styles/fills
    wb_vals   = load_workbook(V4_FILE, data_only=True)
    wb_styles = load_workbook(V4_FILE, data_only=False)

    season_sheets = sorted(
        [s for s in wb_vals.sheetnames if s.startswith('Season')],
        key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 0
    )
    print(f"Seasons found: {[s for s in season_sheets]}")
    if SKIP_SEASONS:
        print(f"Skipping seasons (already in v5): {SKIP_SEASONS}\n")
    if ONLY_SEASON_NUMBER is not None:
        print(f"ONLY_SEASON_NUMBER={ONLY_SEASON_NUMBER} — migrating that season tab only.\n")

    wb_out = Workbook()
    wb_out.remove(wb_out.active)  # remove default blank sheet

    for sheet_name in season_sheets:
        season_num = int(sheet_name.split()[-1])

        if ONLY_SEASON_NUMBER is not None and season_num != ONLY_SEASON_NUMBER:
            print(f"Skipping {sheet_name} (only season {ONLY_SEASON_NUMBER} requested)")
            continue

        if season_num in SKIP_SEASONS:
            print(f"Skipping {sheet_name} (in SKIP_SEASONS)")
            continue

        print(f"Migrating {sheet_name} ...")
        ws_vals   = wb_vals[sheet_name]
        ws_styles = wb_styles[sheet_name]
        ws_out = wb_out.create_sheet(title=sheet_name)

        # Write header row
        ws_out.append(V5_HEADER)
        for cell in ws_out[1]:
            cell.font = Font(bold=True)

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        rows = migrate_season(ws_vals, ws_styles, season_num, wb_styles)
        for row_data, team_color, opponent, team_colors in rows:
            ws_out.append(row_data)
            current_row = ws_out.max_row

            team_fill = PatternFill(patternType='solid', fgColor=team_color) if team_color else None
            opp_rgb   = team_colors.get(opponent.lower()) if opponent else None
            opp_fill  = PatternFill(patternType='solid', fgColor=opp_rgb) if opp_rgb else None

            # Team=2, Player=3 — own team color
            for col in (2, 3):
                cell = ws_out.cell(current_row, col)
                cell.font = Font(bold=True)
                cell.border = border
                if team_fill:
                    cell.fill = team_fill

            # Game 5 winner=13 — fill only when a winner is recorded; color matches that team
            g5_val = row_data[12]
            g5_name = str(g5_val).strip() if g5_val else ""
            g5_cell = ws_out.cell(current_row, 13)
            g5_cell.font = Font(bold=True)
            g5_cell.border = border
            if g5_name:
                win_rgb = _team_color_for_name(team_colors, g5_name)
                if win_rgb:
                    g5_cell.fill = PatternFill(patternType='solid', fgColor=win_rgb)
                else:
                    g5_cell.fill = _NO_FILL
            else:
                g5_cell.fill = _NO_FILL

            # Opponent=16 — opponent's color
            opp_cell = ws_out.cell(current_row, 16)
            opp_cell.font = Font(bold=True)
            opp_cell.border = border
            if opp_fill:
                opp_cell.fill = opp_fill

        print(f"  -> {len(rows)} rows written")

    print(f"\nSaving to {OUTPUT_FILE} ...")
    wb_out.save(OUTPUT_FILE)
    print("Done! Review the output file, then copy any sheets you want into v5.")


if __name__ == "__main__":
    main()
