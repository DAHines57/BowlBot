"""
Script to create an Excel file with database-like structure for bowling league data.
Handles both individual games and aggregated data for older seasons.
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_database_excel(output_file="Bowling_League_Database.xlsx"):
    """Create a new Excel workbook with database-like sheets."""
    
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. PLAYERS Sheet
    ws_players = wb.create_sheet("players", 0)
    headers = ["player_id", "player_name", "created_at", "updated_at"]
    ws_players.append(headers)
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws_players.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add example data row
    ws_players.append([1, "Example Player", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""])
    
    # Set column widths
    ws_players.column_dimensions['A'].width = 12
    ws_players.column_dimensions['B'].width = 30
    ws_players.column_dimensions['C'].width = 20
    ws_players.column_dimensions['D'].width = 20
    
    # Freeze header row
    ws_players.freeze_panes = "A2"
    
    # 2. SEASONS Sheet
    ws_seasons = wb.create_sheet("seasons", 1)
    headers = ["season_id", "season_number", "season_name", "start_date", "end_date", "is_active", "created_at"]
    ws_seasons.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_seasons.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add example seasons
    for i in range(1, 11):
        ws_seasons.append([
            i,
            i,
            f"Season {i}",
            "",  # start_date
            "",  # end_date
            "TRUE" if i == 10 else "FALSE",  # is_active
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    ws_seasons.column_dimensions['A'].width = 12
    ws_seasons.column_dimensions['B'].width = 15
    ws_seasons.column_dimensions['C'].width = 20
    ws_seasons.column_dimensions['D'].width = 12
    ws_seasons.column_dimensions['E'].width = 12
    ws_seasons.column_dimensions['F'].width = 12
    ws_seasons.column_dimensions['G'].width = 20
    ws_seasons.freeze_panes = "A2"
    
    # 3. TEAMS Sheet
    ws_teams = wb.create_sheet("teams", 2)
    headers = ["team_id", "season_id", "team_name", "created_at"]
    ws_teams.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_teams.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add example teams
    example_teams = [
        (1, 10, "Rolling Stoned"),
        (2, 10, "Irregular Bowl Movements"),
        (3, 10, "Gutter Chaos"),
        (4, 10, "Pin-etration Nation"),
    ]
    for team_id, season_id, team_name in example_teams:
        ws_teams.append([team_id, season_id, team_name, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    
    ws_teams.column_dimensions['A'].width = 12
    ws_teams.column_dimensions['B'].width = 12
    ws_teams.column_dimensions['C'].width = 35
    ws_teams.column_dimensions['D'].width = 20
    ws_teams.freeze_panes = "A2"
    
    # 4. TEAM_PLAYERS Sheet (Junction Table)
    ws_team_players = wb.create_sheet("team_players", 3)
    headers = ["team_player_id", "team_id", "player_id", "season_id", "is_captain", "draft_round", "created_at"]
    ws_team_players.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_team_players.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add example relationships
    ws_team_players.append([1, 1, 1, 10, "TRUE", 1, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    
    ws_team_players.column_dimensions['A'].width = 18
    ws_team_players.column_dimensions['B'].width = 12
    ws_team_players.column_dimensions['C'].width = 12
    ws_team_players.column_dimensions['D'].width = 12
    ws_team_players.column_dimensions['E'].width = 12
    ws_team_players.column_dimensions['F'].width = 12
    ws_team_players.column_dimensions['G'].width = 20
    ws_team_players.freeze_panes = "A2"
    
    # 5. GAMES Sheet (Individual Game Scores)
    ws_games = wb.create_sheet("games", 4)
    headers = ["game_id", "player_id", "team_id", "season_id", "week_number", "score", "game_date", "is_playoff", "created_at"]
    ws_games.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_games.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add example games
    ws_games.append([1, 1, 1, 10, 1, 180, "2024-01-15", "FALSE", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_games.append([2, 1, 1, 10, 2, 195, "2024-01-22", "FALSE", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    
    ws_games.column_dimensions['A'].width = 12
    ws_games.column_dimensions['B'].width = 12
    ws_games.column_dimensions['C'].width = 12
    ws_games.column_dimensions['D'].width = 12
    ws_games.column_dimensions['E'].width = 15
    ws_games.column_dimensions['F'].width = 10
    ws_games.column_dimensions['G'].width = 12
    ws_games.column_dimensions['H'].width = 12
    ws_games.column_dimensions['I'].width = 20
    ws_games.freeze_panes = "A2"
    
    # 6. TEAM_STANDINGS Sheet
    ws_standings = wb.create_sheet("team_standings", 5)
    headers = ["standing_id", "team_id", "season_id", "wins", "losses", "ties", "pins_for", "pins_against", 
               "avg_per_game", "avg_per_game_against", "updated_at"]
    ws_standings.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_standings.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add example standings
    ws_standings.append([1, 1, 10, 13, 15, 0, 18500, 18200, 185.2, 182.0, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    
    ws_standings.column_dimensions['A'].width = 15
    ws_standings.column_dimensions['B'].width = 12
    ws_standings.column_dimensions['C'].width = 12
    for col in range(4, 12):
        ws_standings.column_dimensions[get_column_letter(col)].width = 15
    ws_standings.freeze_panes = "A2"
    
    # 7. PLAYER_AGGREGATES Sheet (For seasons without individual games)
    ws_aggregates = wb.create_sheet("player_aggregates", 6)
    headers = ["aggregate_id", "player_id", "team_id", "season_id", "total_games", "total_pins", 
               "average", "high_game", "low_game", "week_1", "week_2", "week_3", "week_4", 
               "week_5", "week_6", "week_7", "notes", "created_at"]
    ws_aggregates.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_aggregates.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add example aggregate data (for older seasons)
    ws_aggregates.append([
        1, 1, 1, 1,  # aggregate_id, player_id, team_id, season_id
        28,  # total_games
        4900,  # total_pins
        175.0,  # average
        220,  # high_game
        120,  # low_game
        180, 185, 170, 190, 175, 180, 185,  # week scores (if available)
        "Aggregated data - individual games not available",  # notes
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])
    
    ws_aggregates.column_dimensions['A'].width = 15
    ws_aggregates.column_dimensions['B'].width = 12
    ws_aggregates.column_dimensions['C'].width = 12
    ws_aggregates.column_dimensions['D'].width = 12
    for col in range(5, 18):
        ws_aggregates.column_dimensions[get_column_letter(col)].width = 12
    ws_aggregates.column_dimensions['Q'].width = 50
    ws_aggregates.column_dimensions['R'].width = 20
    ws_aggregates.freeze_panes = "A2"
    
    # 8. SEASON_CHAMPIONS Sheet
    ws_champions = wb.create_sheet("season_champions", 7)
    headers = ["champion_id", "season_id", "team_id", "notes"]
    ws_champions.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_champions.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Add example champions
    champions_data = [
        (1, 1, 1, "Pin Seekers"),
        (2, 2, 2, "Rolling Stoned"),
        (3, 3, 3, "The Damned"),
    ]
    for champ_id, season_id, team_id, notes in champions_data:
        ws_champions.append([champ_id, season_id, team_id, notes])
    
    ws_champions.column_dimensions['A'].width = 15
    ws_champions.column_dimensions['B'].width = 12
    ws_champions.column_dimensions['C'].width = 12
    ws_champions.column_dimensions['D'].width = 30
    ws_champions.freeze_panes = "A2"
    
    # 9. README Sheet (Instructions)
    ws_readme = wb.create_sheet("README", 0)  # First sheet
    ws_readme.merge_cells('A1:D1')
    title_cell = ws_readme['A1']
    title_cell.value = "Bowling League Database - Excel Template"
    title_cell.font = Font(bold=True, size=16)
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    instructions = [
        "",
        "This Excel file follows a database-like structure for organizing bowling league data.",
        "",
        "SHEETS:",
        "",
        "1. players - Master list of all players",
        "2. seasons - Season information",
        "3. teams - Teams per season",
        "4. team_players - Links players to teams (junction table)",
        "5. games - Individual game scores (for seasons with detailed data)",
        "6. team_standings - Team statistics per season",
        "7. player_aggregates - For older seasons without individual game data",
        "8. season_champions - Season winners",
        "",
        "USAGE:",
        "",
        "• For NEW seasons (Season 10+):",
        "  - Use 'games' sheet for individual game scores",
        "  - Calculate standings from games data",
        "",
        "• For OLD seasons (Season 1-9 or any without individual games):",
        "  - Use 'player_aggregates' sheet",
        "  - Enter total_games, total_pins, average, etc.",
        "  - Optionally fill week columns if you have that data",
        "  - Use 'team_standings' for team-level data",
        "",
        "• Player names should be consistent across all sheets",
        "• Use season_id to link data across sheets",
        "• IDs should be unique and sequential",
        "",
        "MIGRATION FROM OLD EXCEL:",
        "",
        "1. Extract unique player names → players sheet",
        "2. Create season entries → seasons sheet",
        "3. Extract teams per season → teams sheet",
        "4. Link players to teams → team_players sheet",
        "5. For detailed seasons: Convert week columns to games rows",
        "6. For aggregated seasons: Use player_aggregates sheet",
        "7. Copy team standings → team_standings sheet",
        "",
    ]
    
    for i, line in enumerate(instructions, start=2):
        cell = ws_readme.cell(row=i, column=1)
        cell.value = line
        if line and line[0].isupper() and not line.startswith("•"):
            cell.font = Font(bold=True)
    
    ws_readme.column_dimensions['A'].width = 80
    ws_readme.column_dimensions['B'].width = 20
    ws_readme.column_dimensions['C'].width = 20
    ws_readme.column_dimensions['D'].width = 20
    
    # Save workbook
    wb.save(output_file)
    print(f"Created database Excel file: {output_file}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")
    return output_file


if __name__ == "__main__":
    create_database_excel()

